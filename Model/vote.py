import openpyxl
import pandas

from Model import dataFrame_file

from functools import reduce

#Fonction max pour returner le Maximum
def max(x,y):
    if x>=y:
        return x
    else:
        return y


#fonction qui récupére les ligne d un fichiers
def alllignes(file,col=1):
    list=[]
    f=dataFrame_file.openfile(file)
    max_l=f.max_row
    max_c=f.max_column
    for i in range(1,max_l+1):
        list.append(f.cell(row=i,column=col).value)
    return list

#fonction qui retourne la valeur numériqued une ligne

def ligne_value(description,fichier,col=1,col2=2):
    f=dataFrame_file.openfile(fichier)
    for i in range(1,f.max_row+1):
        if description==f.cell(row=i,column=col).value:

            return f.cell(row=i,column=col2).value

    return 0



# fonction qui calcule la somme des votes


def calculer_vote(fichier1,fichier2,col=1,col2=2):

    f1=dataFrame_file.openfile(fichier1)
    f2=dataFrame_file.openfile(fichier2)
    #création d un nouveau workbook
    wb = openpyxl.Workbook()
    f3=wb.active

    #dataFrame_file.create_xlsx_file("resultat_vote.xlsx")

    #file3=openpyxl.load_workbook("resultat_vote.xlsx")
    #f3=file3.active

    lignes=pandas.unique(alllignes(fichier1)+alllignes(fichier2))
    print(alllignes(fichier1))
    print(alllignes(fichier2))
    print (lignes)
    for i in range(len(lignes)):
        #on ecrit la valeur de la premiere colonne dans la ligne i
        f3.cell(row=(i+1),column=col,value=lignes[i])

        #on rajoute la valeur


        if  ligne_value(lignes[i], fichier2)== ligne_value(lignes[i],fichier1) and ligne_value(lignes[i],fichier1)==None:
            f3.cell(row=(i + 1), column=col2, value= 0)

        elif ligne_value(lignes[i], fichier2)== None:
            f3.cell(row=(i + 1), column=col2, value=ligne_value(lignes[i], fichier1) + 0)
        elif ligne_value(lignes[i],fichier1)==None:
            f3.cell(row=(i + 1), column=col2, value=  ligne_value(lignes[i], fichier2))
        else:
            f3.cell(row=(i+1),column=col2,value=ligne_value(lignes[i],fichier1)+ligne_value(lignes[i],fichier2))



    #------> sauvgarder le ichier File 3
    wb.save("resultat_vote.xlsx")
    return "resultat_vote.xlsx"

#reduce votes
def calculer_all_votes(fichiers):
    return reduce(lambda x,y:calculer_vote(x,y),fichiers)




