import openpyxl

#--------------------------------------------
#------detecter le nom de la fonction -------
#--------------------------------------------

def get_fucntion_name(str):
    if str.startswith("="):
        return str.split("(")[0][1:]
    else:
        return None


import re

#--------------------------------------------
#------Cellule ou Fonction-------
#--------------------------------------------

def is_xlsx_cell(s):
    # definer une expression reguliére
    pattern = r'^[A-Z]+\d+$'

    # regarder si le characére matsh avec l expression réguliére
    if re.match(pattern, s):
        return True
    else:
        return False

#--------------------------------------------
#------Formules ?-------
#--------------------------------------------

def is_xlsx_formula(s):

    pattern = r'^\=[A-Z]+\([A-Za-z0-9\,\+\-\*\/\(\)]*\)$'


    if re.match(pattern, s):
        return True
    else:
        return False


#----------------------------------------
#---------CONTENU DU COLONNE-------------
#----------------------------------------

def get_column_values(column_number, file_name):

    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook.active
    column_values = []

    for cell in worksheet.iter_cols(min_col=column_number, max_col=column_number):
        for cell_value in cell:
            column_values.append(cell_value.value)

    return column_values
#----------------------------------------
#----------CONTENU DU LIGNE--------------
#----------------------------------------



def get_ligne(ligne_number, file_name):
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook.active
    ligne_values = []

    for cell in worksheet[ligne_number]:
        ligne_values.append(cell.value)

    return ligne_values
#----------------------------------------
#------detecter une formule--------------
#----------------------------------------
def est_formule(chaine):
    if "=" in chaine:
        return True
    else:
        return False
