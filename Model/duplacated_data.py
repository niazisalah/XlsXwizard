from Model import combine
from Model import vote
import openpyxl
import pandas

def clean_dup(file_path):
    # lire le fichier a l'aide de pandas
    df = pandas.read_excel(file_path)

    # Compter le nombre des duplicatas
    duplicats = df.duplicated(keep=False).sum()

    # If there are no duplicates, return the dataframe as is
    if duplicats == 0:
        return df

    # Add a new column to the dataframe to store the count of duplicates
    df['NOMBRE DE DUBLICATAS'] = df.duplicated(keep='first').astype(int)

    # Drop all the duplicates
    df.drop_duplicates(keep='first', inplace=True)

    # Save the modified dataframe to a new xlsx file
    df.to_excel(file_path[:-5] + '_modifié.xlsx', index=False)

    # Return the modified dataframe
    return df

def compare_ligne(l1,l2):
    comparaison = list(map(lambda i, j: i == j, l1, l2))
    comparaison=list( filter (lambda x:x==False,comparaison))
    return len(comparaison)

def compter_duplicats(l,fichier):
    f = openpyxl.load_workbook(fichier)
    fw = f.active
    max_l = fw.max_row
    max_c = fw.max_column
    liste=[]
    nbr_duplicat=0
    for i in range (1,max_l+1):
        for j in range(1, max_c + 1):
            liste.append(fw.cell(row=i,column=j).value)
        if(compare_ligne(l,liste)==0):
            nbr_duplicat =nbr_duplicat+1
            print(liste)
            liste=[]
            print(liste)
            if(nbr_duplicat>1):
                delete_line(fichier, i)
    if(nbr_duplicat==0):
        return 0
    else:
        return (nbr_duplicat - 1)

def duplcate_count(file):
    f=openpyxl.load_workbook(file)
    fw=f.active
    max_l=fw.max_row
    max_c=fw.max_column
    list1=[]
    list2=[]
    list3=[]
    for i in range(1,max_l+1):
        for j in range (1,max_c+1):
            list1.append(fw.cell(row=i,column=j).value)
        var=compter_duplicats(list1,file)

        if(var!=0):
            list2.append(i)
            list3.append(var)

    return list(zip(list2,list3))

def duplicate_write(fichier,list):
    f = openpyxl.load_workbook(fichier)
    fw = f.active

    maxc=fw.max_column+1
    for l in list:
        fw.cell(row=l[0],column=maxc,value=l[1])
    f.save(fichier)
    return fichier

def traiter_duplicats(fichier):
    duplcate_count(fichier)
    duplicate_write(fichier,duplcate_count(fichier))


#a refaire
def ecrir_valeurs(file,list):
    f = openpyxl.load_workbook(file)
    fw = f.active
    l=[]
    for i in range(list):
        for j in range(1,fw.max_row+1):
            for k in range(1,fw.max_column+1):
                #non complet
                l.append(fw.cell(row=j,column=k,value=list).value)
            if (compare_ligne(l,list)==0):
                #nombre de différence egale a 0 donc c'est la bonne ligne
                fw.cell(row=j,column=fw.max_column+1,value=list[len(list)-1])


#supprimer les lignes dupliqué
def remove_duplicate_rows(file_path):
    # Read the XLSX file into a pandas dataframe
    df = pandas.read_excel(file_path)

    # Remove duplicate rows from the dataframe
    df.drop_duplicates(inplace=True)

    # Write the updated dataframe back to the XLSX file
    df.to_excel(file_path, index=False)



def delete_line(file_path, line_number):
    # Load the xlsx file
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Delete the specified line
    sheet.delete_rows(line_number, 1)

    # Save the changes
    wb.save(file_path)
