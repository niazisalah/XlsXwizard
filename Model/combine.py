import pandas
from functools import reduce
from openpyxl import load_workbook
from ProjetSynthese.Model import dataFrame_file


#---------------------------------------------------------------------------------
#------------------------------Combinaison vertical-----------------------------
#---------------------------------------------------------------------------------

def combiner_datahseet_v(df1, df2):
    # lire le premier fichier

    #f1 = pandas.read_excel(fichier1)

    # Lire le deuxiéme fichier
    #f2 = pandas.read_excel(fichier2)

    # combiner les deux fichier a l'aide de concat
    return pandas.concat([df1, df2],axis=0)



def gocomb(f1,f2):
    wbtoxlsx(combiner(filetowb(f1),filetowb(f2)))
def filetowb(file):
    wb = load_workbook(file)

    return wb

def combiner(wb1, wb2):

    # Charger les données du fichier 1

    feuille1 = wb1.active

    # Charger les données du fichier 2

    feuille2 = wb2.active

    # Copier les données de la feuille 2 à la fin de la feuille 1
    for row in feuille2.iter_rows(values_only=True):
        feuille1.append(row)

    # Enregistrer le résultat dans un nouveau fichier Excel
    #wb1.save(fichier_sortie)
    return wb1

def wbtoxlsx(wb, fichier_sortie="result.xlsx"):
    wb.save(fichier_sortie)


    return fichier_sortie








#---------------------------------------------------------------------------------
#----------------------------Combinaison horizentale------------------------------
#---------------------------------------------------------------------------------

def combiner_dataframe_h(fichier1,fichier2):
    # Charger le premier fichier
    df1 = pandas.read_excel(fichier1)

    # Charger le deuxiéme fichier
    df2 = pandas.read_excel(fichier2)

    # Determiner le nombre de lignes dans chaque fichier
    rows_df1 = df1.shape[0]
    rows_df2 = df2.shape[0]

    
    if rows_df1 > rows_df2:
        padding = pandas.DataFrame(index=range(rows_df1 - rows_df2), columns=df2.columns)
        df2 = pandas.concat([df2, padding], ignore_index=True)
    elif rows_df2 > rows_df1:
        padding = pandas.DataFrame(index=range(rows_df2 - rows_df1), columns=df1.columns)
        df1 = pandas.concat([df1, padding], ignore_index=True)

    # Combiner deux  dataframes horizentalement (column-wise)


    return pandas.concat([df1, df2], axis=1)

#---------------------------------------------------------------------------------
#---------Prend une liste de fichier et retourne une liste de DataFrame-----------
#---------------------------------------------------------------------------------

def files_towbs(fichiers):
    return list(map(filetowb, fichiers))

#---------------------------------------------------------------------------------
#---------prend deux valeur en entré une fonction et une liste de fichier---------
#---------------------------------------------------------------------------------

def combiner_tout(fichiers):

    wbs =files_towbs(fichiers)
    DataFrame=reduce(lambda x,y:combiner(x,y),wbs)
    print("combinaison en cours")
    print(wbtoxlsx(DataFrame))
    return wbtoxlsx(DataFrame)
def combiner_tout2(fichiers):

    #DataFrames =files_toDataFrames(fichiers)
    DataFrame=reduce(lambda x,y:combiner_dataframe_h(x,y),fichiers)

    return dataFrame_file.dataFrame_tofile(DataFrame)

def files_toDataFrames(fichiers):
    return list(map(dataFrame_file.file_toDataFrame, fichiers))