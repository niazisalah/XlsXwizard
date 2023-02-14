import openpyxl

class file:
    #intialize the file name
    def __int__(self,name):
        self.name=name

    #read file
    def readfile(self,path):
        wb= openpyxl.load_workbook('emploidutemps.xlsx')
        sheets=wb.sheetnames
        return wb

    def createfile(self,name):
        print("fichié crée")

    def create_xlsx_file(file_name):
        wb = openpyxl.Workbook()
        ws = wb.active
        wb.save(file_name)
        return file_name
