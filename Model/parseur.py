import pandas
from ProjetSynthese.Model import analyseur_cellules, combine, dataFrame_file


def parseur(fichier_specification):
    spec= dataFrame_file.openfile(fichier_specification)

    #on determine la dimension du fichier de specification

    max_l=spec.max_row
    max_c=spec.max_column

    #on va determiner les bloques de fichiers
    # sous forme Colomne de début et colomne de fin ,ligne début, ligne de fin

    file1_spec=[1,0,1,0]
    file2_spec=[0,0,1,0]
    file_result_spec=[0,0,1,0]


#remplir les données sur les colomnes
    i=1
    while spec.cell(row=1,column=i).value!=None:
        i=i+1

    file1_spec[1] = i-1
    i = i + 1
    #ON DECALE une colomne vide entre chaque specification

    #On décale de deux une colomn vide et on pass a la prochaine
    i=file1_spec[1]+2
    # on retrouve la premiére colomn qui concerne le fichier 2
    file2_spec[0] = i

    #on avance petit a petit
    while spec.cell(row=1,column=i+1).value!=None:
        i=i+1
    file2_spec[1]=i
    file_result_spec[0] = i+2
    i=i+1

    #on continue d avancer
    while spec.cell(row=1,column=i+1).value!=None:
        i=i+1
    file_result_spec[1] = i

    #on determine le nombre de lignes
    j = 1
    while spec.cell(row=j, column=1).value != None:
        j = j + 1

    file1_spec[3] = j-1
    j = 1


    # Maintena la spécification du fichier 2
    while spec.cell(row=j, column=file2_spec[0]).value != None:
        j = j + 1
    file2_spec[3] = j-1

    j =  1

    # on continue d avancer
    while spec.cell(row=j, column=file_result_spec[0]).value != None:
        j = j + 1
    file_result_spec[3] = j-1




#la structure suivant va permettre de definir les limites
    return [file1_spec,file2_spec,file_result_spec]

#------------------------------------------------------------
#fonction pour determiner les dimensions des specifications
#------------------------------------------------------------

def dimension(list):
    l = [] #lignes
    c=[] #colomnes
    c.append(list[0][1] - list[0][0])
    c.append(list[1][1] - list[1][0])
    c.append(list[2][1] - list[2][0])

    l.append(list[0][3] - list[0][2])
    l.append(list[1][3] - list[1][2])
    l.append(list[2][3] - list[2][2])


    return l+c
#----------------------------------------
#-------comparer deux listes-------------
#----------------------------------------

def compare_listes(liste1, liste2):
    if len(liste1) != len(liste2):
        return False
    for element in liste1:
        if element not in liste2:
            return False
    return True

#----------------------------------------
#ajouter les elements de deux listes-----
#----------------------------------------

def add_lists(list1, list2):
    # Vérifie que les deux listes ont la même longueur
    if len(list1) != len(list2):
        raise ValueError("Les deux listes doivent avoir la même longueur")

    # Initialise une nouvelle liste pour stocker le résultat
    result = []

    # Parcourt les listes et additionne les éléments correspondants
    for i in range(len(list1)):
        sum = list1[i] + list2[i]
        result.append(sum)

    return result

#----------------------------------------
#-----------les detecteurs---------------
#----------------------------------------

def dict_to_excel(dict_obj):
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write the keys to the first row of the worksheet
    headers = list(dict_obj.keys())
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write the values to subsequent rows
    for row_num, values in enumerate(zip(*dict_obj.values()), 2):
        for col_num, value in enumerate(values, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Save the workbook to a file
    #wb.save(filename)
    return wb





def compare_workbooks(wb1, wb2):
    # Load the workbooks


    # Loop through each worksheet in each workbook and compare the cell values
    for sheet1, sheet2 in zip(wb1.worksheets, wb2.worksheets):
        # Compare the dimensions of the worksheets
        if sheet1.max_row != sheet2.max_row or sheet1.max_column != sheet2.max_column:
            print(f"Sheet {sheet1.title} has different dimensions")
            continue

        # Compare each cell value in the worksheets
        for row in range(1, sheet1.max_row + 1):
            for col in range(1, sheet1.max_column + 1):
                cell1 = sheet1.cell(row=row, column=col)
                cell2 = sheet2.cell(row=row, column=col)
                if cell1.value != cell2.value:
                    print(f"Sheet {sheet1.title}, cell {cell1.coordinate}: {cell1.value} != {cell2.value}")


def determiner_fonction(fichier_specification):
    #a l interieur on va determiner s il ya une spécification et retourner

    list=parseur(fichier_specification)
    file1=  xlsx_toKV(fichier_specification,  list[0][1])
    file2 = xlsx_toKV(fichier_specification, list[1][1])
    file3 = xlsx_toKV(fichier_specification, list[2][1])
    wb1 = dict_to_excel(file1)
    wb2 = dict_to_excel(file2)
    wb3 = dict_to_excel(file3)

    #on test vote

    #on test la combinaison vertical

    #on test la combinaison horizental



    return 0

#----------------------------------------
#----------------------------------------

def try_combinaison_vertical(fichier_specification):

    list = parseur(fichier_specification)
    dim = dimension(list)

    #premiére condition
    print("on essaye la combinaison vertical")
    if dim[2] == dim[1] + dim[0]:
        print("on essaye la combinaison vertical")
        # deuxiéme condition faut vérifier si le conetenu de colomne est le meme

        return True
    else:
        return False

#----------------------------------------
#----------------------------------------

def try_combinaison_horizental(fichier_specification):
    list = parseur(fichier_specification)
    dim = dimension(list)
    print("on essaye la combinaison horizetal")
    #on test si la dimension le nombre de lignes
    if dim[5] > dim[4] and dim[5] > dim[3]:
        print("on essaye la combinaison horizetal")
        return True
    else:
        return False

#----------------------------------------
#----------------------------------------

def detecter_xlsx_fonctions(fichier_specification):
    #->lire le fichier
    list = parseur(fichier_specification)
    #->on recupére le contenu de la colomne de fichier 3 dans le fichier de spécification
    fonctions = analyseur_cellules.get_column_values(list[2][1], fichier_specification)

    #on essaye de voir si la colomne de spécification contient des formules ou non
    for fonction in fonctions:
        if analyseur_cellules.is_xlsx_formula(fonction):
            return False
    return True




#----------------------------------------
#-----Fonction pour parser le vote-------
#----------------------------------------

def try_vote(fichier_specification):


    list = parseur(fichier_specification)

    dataspec= analyseur_cellules.get_column_values(list[2][0], fichier_specification)
    datafile1 = analyseur_cellules.get_column_values(list[0][0], fichier_specification)
    datafile2 = analyseur_cellules.get_column_values(list[1][0], fichier_specification)


    if compare_listes(pandas.unique(datafile1+datafile2),dataspec):
        dataspec = analyseur_cellules.get_column_values(list[2][1], fichier_specification)
        datafile1 = analyseur_cellules.get_column_values(list[0][1], fichier_specification)
        datafile2 = analyseur_cellules.get_column_values(list[1][1], fichier_specification)

        if compare_listes(dataspec,add_lists(datafile1,datafile2)):
            return True
        else:
            return False

    else:
        return False


#----------------------------------------
#------passage au key value--------------
#----------------------------------------

def xlsx_toKV(file,colomne):
    l= parseur(file)
    dict={}
    spec = dataFrame_file.openfile(file)
    values= analyseur_cellules.get_column_values(colomne, file)
    keys = analyseur_cellules.get_column_values(colomne - 1, file)

    for i in range (values):
        dict[keys[i]]=values[i]

    return dict



def try_dublicats(fichier_specification):
    return True

def correction(fichier_specification):
    return True

import openpyxl

def compare_xlsx_files(file1, file2):

    # Load the Excel files
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    # Iterate through all sheets in both files
    for sheetname in wb1.sheetnames:
        if sheetname not in wb2.sheetnames:
            # If sheetname not found in wb2, return False
            return False

        # Get the corresponding sheet from wb2
        sheet1 = wb1[sheetname]
        sheet2 = wb2[sheetname]

        # Iterate through all cells in the sheet
        for row in range(1, sheet1.max_row + 1):
            for col in range(1, sheet1.max_column + 1):
                # Compare the cell values
                if sheet1.cell(row=row, column=col).value != sheet2.cell(row=row, column=col).value:
                    # If cell values are not equal, return False
                    return False

    # If all cells in all sheets are equal, return True
    return True


def extraire_donnees(nom_fichier, col_debut, col_fin, ligne_debut, ligne_fin, nom_nouveau_fichier):
    # Charger le fichier xlsx
    wb = openpyxl.load_workbook(nom_fichier)

    # Sélectionner la feuille de calcul active
    feuille = wb.active

    # Initialiser la liste pour stocker les données extraites
    donnees = []

    # Parcourir les lignes spécifiées
    for i in range(ligne_debut, ligne_fin+1):
        # Initialiser la liste pour stocker les données de la ligne courante
        ligne = []
        # Parcourir les colonnes spécifiées
        for j in range(col_debut, col_fin+1):
            # Ajouter la valeur de la cellule courante à la liste de la ligne
            ligne.append(feuille.cell(row=i, column=j).value)
        # Ajouter la liste de la ligne à la liste de données
        donnees.append(ligne)

    # Créer un nouveau fichier xlsx pour stocker les données extraites
    nouveau_wb = openpyxl.Workbook()
    nouvelle_feuille = nouveau_wb.active

    # Écrire les données extraites dans la nouvelle feuille de calcul
    for i in range(len(donnees)):
        for j in range(len(donnees[i])):
            nouvelle_feuille.cell(row=i+1, column=j+1).value = donnees[i][j]

    # Enregistrer le nouveau fichier xlsx
    nouveau_wb.save(nom_nouveau_fichier)


def trycombh(template):
    list=parseur(template)

    extraire_donnees(template, list[0][0], list[0][1], list[0][2], list[0][3], "temp1.xlsx")
    extraire_donnees(template, list[1][0], list[1][1], list[1][2], list[1][3], "temp2.xlsx")
    extraire_donnees(template, list[2][0], list[2][1], list[2][2], list[2][3], "temp3.xlsx")
    t= combine.combiner_tout(["temp1.xlsx", "temp2.xlsx"])

    if compare_xlsx_files(t,"temp3.xlsx"):
        return True
    else:
        return False

def trycombv(template):
    list=parseur(template)
    print(list)
    extraire_donnees(template, list[0][0], list[0][1], list[0][2], list[0][3], "temp1.xlsx")
    extraire_donnees(template, list[1][0], list[1][1], list[1][2], list[1][3], "temp2.xlsx")
    extraire_donnees(template, list[2][0], list[2][1], list[2][2], list[2][3], "temp3.xlsx")
    t= combine.combiner_tout2(["temp1.xlsx", "temp2.xlsx"])
    print(t)
    if compare_xlsx_files(t,"temp3.xlsx"):
        return True
    else:
        return False
