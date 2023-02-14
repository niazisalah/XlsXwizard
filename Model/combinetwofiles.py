import openpyxl
from Model import comparer_ligne

def openfile(fichier):
    file=openpyxl.load_workbook(fichier)
    return file.active

def combiner(fichier1,fichier2):
# on charge le fichier 1
    file1 = openpyxl.load_workbook(fichier1)
    file1_sheet = file1.active

# on charge le fichier 2
    file2 = openpyxl.load_workbook(fichier2)
    file2_sheet = file2.active

# Récuperer le  maximum ligne et  colomne du fichier 2
    fichier2_max_ligne = file2_sheet.max_row
    fichier2_max_colomne = file2_sheet.max_column

# Copier les données du fichier 2 ET Les ajouter dans le fichier 1 pour pouvoir utiliser un reduce en fin
# a faire (NON COMPLET)
# si les colomnes du fichier 2 son inferieur ou égale (sans prendre le cas ou les colonme ont des nom différents)
    if file2_sheet.max_column <= file1_sheet.max_column:
        # il faut proceder a la comparaison des entêtes
        entete1=comparer_ligne.recuperer_entete(fichier1)
        entete2=comparer_ligne.recuperer_entete(fichier2)
        for row in range(1, fichier2_max_ligne + 1):
            ligne_courante = file1_sheet.max_row + 1
        # on ignore l entete du fichier 2 car c la meme que celle du fichier 1
            for col in range(1, fichier2_max_colomne + 1):
                file1_sheet.cell(row=ligne_courante, column=col, value=file2_sheet.cell(row=row, column=col).value)
    else:
    # On calcule la diférence pour ajouter les nouvelle colomne dans l'entête du fichier

        diference = file2_sheet.max_column - file1_sheet.max_column
        s = file1_sheet.max_column + 1
        for i in range(diference):
            file1_sheet.cell(row=1, column=s, value=file2_sheet.cell(row=1, column=s).value)
            s = s + 1

        for row in range(2, fichier2_max_ligne + 1):
            ligne_courante = file1_sheet.max_row + 1
            for col in range(1, fichier2_max_colomne + 1):
                file1_sheet.cell(row=ligne_courante, column=col, value=file2_sheet.cell(row=row, column=col).value)

# on enregistre la combinason dans le fihcier1
    file1.save(fichier1)
    print("combinason réussi")
