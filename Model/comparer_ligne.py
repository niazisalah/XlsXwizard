import openpyxl
#On va avoir besoin de cette fonction lorsqu on fusionner deux fichier avec des entêtes différentes
def comparer_ligne(liste1,liste2):
    #LISTE 1 ET 2 DOIVENT AVOIR LA MËME TAILLE
    comparaison=list(map(lambda i,j:i==j,liste1,liste2))
    return comparaison

def detecter_la_comparason(liste):
    return all(liste)

#fonction pour recuperer
def recuperer_entete(fichier):
    file = openpyxl.load_workbook(fichier)
    file_sheet = file.active
    entete=[]
    for col in range(1,file_sheet.max_row+1):
        entete.append(file_sheet.cell(row=1, column=col).value)
    return entete



