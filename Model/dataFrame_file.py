import openpyxl
import pandas


#---------------------------------------------------------------------------------
#---------------------Fonction pour lire le fichier xlsx--------------------------
#---------------------------------------------------------------------------------
def openfile(fichier):
    file=openpyxl.load_workbook(fichier)
    return file.active
#---------------------------------------------------------------------------------
#----------------------------Creation d'un fichier xlsx---------------------------
#---------------------------------------------------------------------------------

def create_xlsx_file(file_name):

    wb = openpyxl.Workbook()
    ws = wb.active
    wb.save(file_name)

#---------------------------------------------------------------------------------
#--------------------La fonction qui convertie la feuille en fichier--------------
#---------------------------------------------------------------------------------
def dataFrame_tofile(DataFrame):
    create_xlsx_file("result_.xlsx")
    DataFrame.to_excel("result_.xlsx", index=False)
    disable_bold_border("result_.xlsx")
    return "result.xlsx"

#---------------------------------------------------------------------------------
#--------------fonction qui convertis un fichier xlsx to DataFrame----------------
#---------------------------------------------------------------------------------

def file_toDataFrame(fichier):
    ds= pandas.read_excel(fichier)
    return ds
#---------------------------------------------------------------------------------
#-----------fonction pour désactiver les border et le bold de l'entête------------
#---------------------------------------------------------------------------------

def disable_bold_border(fichier):

    wb = openpyxl.load_workbook(fichier)
    sheet = wb.active


    for row in sheet.iter_rows():
        for cell in row:
            cell.font = openpyxl.styles.Font(bold=False)
            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style=None),
                                                right=openpyxl.styles.Side(border_style=None),
                                                top=openpyxl.styles.Side(border_style=None),
                                                bottom=openpyxl.styles.Side(border_style=None))


    wb.save(fichier)